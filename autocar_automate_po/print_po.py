try:
    import openpyxl as pyxl
    from openpyxl.styles import Alignment
except ImportError:
    print('Please install the following dependencies:')
    print('openpyxl')
    _ = input('Press Enter to exit\n')
    exit()

# Standard
from pathlib import Path
from datetime import datetime
import os
from time import sleep


def main():
    # Declare variables
    files_not_opened = []
    # Get unique identifier
    time_identifier = datetime.now().strftime('%m%d%y-%H%M%S')
    i_identifier = 1

    Path('PR_Processed_Files').mkdir(exist_ok=True)

    for path in Path('Files_To_Process').glob('*.xlsx'):
        try:
            po_identifier = 'PR-' + time_identifier + '-' + str(i_identifier)
            change_file(path, po_identifier)
            print('Printing file: ' + po_identifier)
            os.startfile(Path('PR_Processed_Files') / (po_identifier + '.xlsx'), 'print')
            print('Loading...')
            sleep(10)
            i_identifier += 1
        except Exception:
            files_not_opened.append(path.stem)

    if files_not_opened:
        print('Files not opened:')
        print(*files_not_opened, sep='\n')

    # ----------------------------------------------------------------------------------------------------
    print('Complete')
    _ = input('Press Enter to exit\n')


def change_file(path, po_identifier):
    wb = pyxl.load_workbook(filename=path, data_only=True)
    ws = wb.active

    ws['C1'].value = po_identifier
    ws['C1'].alignment = Alignment(wrap_text=False)
    wb.save(Path('Processed_Files') / (po_identifier + '.xlsx'))


if __name__ == '__main__':
    main()
